"""Generate Docs/Prasad_TrackA_Task_Tracker.xlsx — Prasad's (Track A) task list
and work log. Mirrors the 3-sheet format of Shagun_Task_Tracker.xlsx.

Regenerate with:  python build_prasad_tracker.py

Sheet 1 "Task List"  — every Track A / Sprint-0 task, its POA reference, status,
                       completion date and a "What I did" column filled in as
                       each task lands.
Sheet 2 "Work Log"   — dated, one row per session/commit, so the tracker shows
                       *what was actually done* rather than only a status flag.
Sheet 3 "Reference"  — track split, shared-file rules, decisions, open questions.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "Docs/Prasad_TrackA_Task_Tracker.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
PROG_FILL = PatternFill("solid", fgColor="FFEB9C")
TODO_FILL = PatternFill("solid", fgColor="F2F2F2")
BLOCK_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "DONE": DONE_FILL,
    "In progress": PROG_FILL,
    "Not started": TODO_FILL,
    "Blocked": BLOCK_FILL,
}

# --------------------------------------------------------------------------- #
# Sheet 1 — Task List
# --------------------------------------------------------------------------- #
TASK_COLUMNS = [
    ("Task ID", 10),
    ("Task", 46),
    ("POA reference", 24),
    ("Phase", 14),
    ("Status", 13),
    ("Started", 12),
    ("Completed", 12),
    ("What I did (fill in on completion)", 74),
    ("Files / evidence", 46),
]

TASKS = [
    # --- Sprint 0: shared test-dataset backlog (POA/16 §16) ----------------- #
    ("S1", "Expand taxonomy & stations — 12 classes, city/suburban, one-way",
     "POA/16 §16.2", "Sprint 0", "DONE", "2026-09-01", "2026-09-01",
     "Expanded to the full 12-class Hertz-style ACRISS taxonomy and added city/suburban stations "
     "plus a new US/USD region and domestic one-way support. KEY DECISION: added the 8 new classes "
     "and 6 new stations in a purely ADDITIVE world pass-3, so passes 1 & 2 keep their exact RNG "
     "draw sequence and the 7 golden-scenario prices stay byte-identical — proven by recomputing "
     "them against origin/main in a throwaway worktree (LHR/ICAR=52.21, MAN/ECAR=31.47, "
     "FRA/FCAR=60.63). Kept the legacy generic 'SUV' code for backward-compat with the existing "
     "golden fixtures rather than renaming it (a rename would shift the RNG and move the prices). "
     "One-way is domestic-only (cross-border needs approval per the policy). Rewrote the "
     "per-location currency test to be country-driven so future stations don't break it.",
     "generator/world.py, catalogues.py, config.py, models.py (LocationType.suburban); "
     "tests/test_taxonomy_stations.py (8). Commit 73dde84 (pushed track-a, PR #1)"),

    ("S2", "Fee completeness — late-return, no-show, fuel + dispute scenarios",
     "POA/16 §16.1", "Sprint 0", "DONE", "2026-09-01", "2026-09-01",
     "Added itemised, disputable fees. KEY DECISION: post-rental charges (late-return, no-show, "
     "fuel) are recorded as separate FeeLine rows and deliberately NOT folded into the quoted "
     "total, so 'why was I charged more than I was quoted?' disputes are realistic; the one-way "
     "fee, which IS known at booking, does go into the total. Modelled no-show as a first-class "
     "BookingStatus. Put the fee amounts in a single fee_rules() source of truth so generated "
     "bookings, the dispute fixtures and the policy text can't drift apart. Shipped 5 hand-authored "
     "dispute fixtures (FeeDisputeComposer) with pinned resolutions covering upheld / refunded / "
     "partial-refund / escalated-to-human, each grounded in the same world the bot verifies against.",
     "generator/customers.py, durations.py (late_return_extra_days), scenarios.py "
     "(FeeDisputeComposer), catalogues.py (fee_rules), models.py (FeeLine/FeeType/FeeDispute/"
     "DisputeResolution, Booking.one_way_fee/fees, BookingStatus.no_show); "
     "tests/test_fees_and_disputes.py (7). Commit 73dde84"),

    ("S5", "Load / SLA config knobs — eps, concurrency, mock timeouts",
     "POA/16 §16.3", "Sprint 0", "DONE", "2026-09-01", "2026-09-01",
     "Turned the resolved §16.3 load/SLA numbers into explicit config knobs rather than magic "
     "numbers: a LoadProfile on GenConfig (10/50/100 eps; 500/1000 concurrent; <2s no-call / <5s "
     "tool-backed SLA; 5s deterministic mock timeout), exposed via VolumeSampler.load_summary and "
     "written to config/load_profile.yaml so the soak/load harness reads one source of truth.",
     "generator/config.py (LoadProfile), volume.py, pipeline.py, models.py (Dataset.load_profile); "
     "tests/test_load_config.py (4). Commit 73dde84"),

    # --- Track A: service modules (POA 01-07, 15) --------------------------- #
    ("A1", "Platform skeleton — repo layout, CI, secrets, logging/tracing baseline",
     "POA/15", "Track A", "DONE", "2026-09-01", "2026-09-01",
     "Started Track A service code. Ratified the services/ monorepo layout (resolves POA/18 §8.2, "
     "recorded in POA/15 §12) and built the shared services/platform template: create_app(name) "
     "returns a FastAPI app with structured JSON logging + per-request correlation id, Prometheus "
     "/metrics, an OpenTelemetry tracing SEAM (off unless HFB_OTEL_ENABLED + libs present, so no "
     "hard OTel dep yet), /healthz + /readyz with pluggable readiness checks, and problem+json "
     "error handling. Data seams (Postgres/Redis/Celery) are lazy factories, so the skeleton needs "
     "none of those deps installed — A2/A7 add them. services/event_pipeline boots on the template "
     "(A2-A8 mount here). Env-driven HFB_* config, no secrets in code.",
     "services/platform/*, services/event_pipeline/*, docker-compose.yml, Dockerfile, "
     ".github/workflows/ci.yml (pytest matrix + ruff), .env.example, POA/15 §12; "
     "tests/test_platform_skeleton.py (8) — 100 total green, ruff clean. Commit 4a1ed3d (PR #1)"),

    ("A2", "Event Store — Postgres + Redis Streams",
     "POA/03", "Track A", "In progress", "2026-09-01", "",
     "Core done + downstream-ready. Built services/event_pipeline/store/: append-only events table "
     "+ transactional event_outbox (POA/03 §3.1); idempotent write_event that inserts both in ONE "
     "transaction (existence check + IntegrityError backstop, never double-enqueues); an OutboxRelay "
     "that publishes THEN marks per-row (at-least-once, consumers dedupe on event_id) — property-"
     "tested for no-loss under a flaky publisher; a Redis StreamPublisher (XADD events:in + "
     "ensure_group trigger-eval) behind a protocol so it's tested with no Redis; read models "
     "recent_events/session_events/last_event_at/has_repeated_search (signal I). Wired a postgres "
     "readiness check + app.state.event_store (A4 writes through it). Portable schema runs on "
     "Postgres (JSONB, bigserial) in prod and in-memory SQLite in tests. DEFERRED (POA/03 §11): "
     "retention/partition job, signal-J booking backfill, live-Redis integration; sync (not async) "
     "+ create_all bootstrap instead of Alembic, both by design.",
     "services/event_pipeline/store/{tables,store,publisher,relay,bootstrap}.py, main.py; "
     "services/requirements.txt, pyproject, POA/03 §11; tests/test_event_store.py (12) — 112 total "
     "green, ruff clean. Commit d8fc10b (pushed track-a, PR #1)"),

    ("A6", "Frequency Cap & Precedence Engine",
     "POA/05", "Track A", "In progress", "2026-09-01", "",
     "Core done + tested. services/event_pipeline/frequency/: Postgres engagement ledger (source of "
     "truth; only reserved/confirmed count, rolled_back ignored), sliding-window caps that DELEGATE "
     "to reference.would_fire (so the service and the invariant suite test the same rule), per-"
     "customer global cap + ISO-8601 cooldown, deterministic precedence (weight -> match specificity "
     "-> most-recent signal -> lowest id, order-independent), atomic reserve under a per-customer "
     "lock (concurrency test: 10 threads, cap 1 -> exactly 1 fires). reserve->confirm/rollback so a "
     "failed M08 send frees the slot instead of burning the cap. Every suppression carries a reason "
     "+ per-trigger losers for M14. CROSS-TRACK: A6->M08 contract (MatchCandidate/EngagementDecision/"
     "SuppressionReason) added to generator/models.py — reconcile with Shagun's local §5 handshake. "
     "Deferred (POA/05 §11): Redis counters, reservation-TTL sweep, M13 cap config. No HTTP surface "
     "(library called by A5).",
     "services/event_pipeline/frequency/{tables,ledger,engine,precedence,lock,bootstrap}.py, "
     "generator/models.py (contract), POA/05 §11; tests/test_frequency_precedence.py (11) — 134 "
     "total green, ruff clean. Commit 7a5a0f8 (PR #1)"),

    ("A4", "Event Ingestion API (FastAPI)",
     "POA/02", "Track A", "In progress", "2026-09-01", "",
     "Core done + tested. The single write-door in services/event_pipeline/ingestion/: POST "
     "/v1/events + POST /v1/events:batch (partial-success) writing through A2's transactional "
     "outbox. Single-event body IS generator.models.Event, so FastAPI validation + extra='forbid' "
     "is the PII field allow-list (POA/15 §4) — unknown/PII field -> 422. Auth, identity binding "
     "and rate limiting sit behind seams because §10 is open: ApiKeyAuthenticator now (mTLS/JWT "
     "later; JWT would set Principal.customer_id so identity binding becomes a real cross-check), "
     "in-memory fixed-window rate limiter (Redis in prod). Idempotency delegated to the store "
     "(retry -> 202 duplicate, never a second row). Outcome->status 202/409/429/503/422. "
     "build_app wires it when a store is present + accepts an injected store for tests. "
     "Deferred (POA/02 §11): real auth mechanism, distributed rate limiter, load verification.",
     "services/event_pipeline/ingestion/{router,service,auth,ratelimit,schemas}.py, main.py, "
     "services/platform/config.py, pyproject (ruff), POA/02 §11; tests/test_ingestion_api.py (11, "
     "incl. e2e API->store->relay->stream) — 123 total green, ruff clean. Commit ee06657 (PR #1)"),

    ("A5", "Trigger Evaluation Engine",
     "POA/04", "Track A", "In progress", "2026-09-01", "",
     "Core done + tested. The node-N brain in services/event_pipeline/triggers/. Sandboxed "
     "field/op/value rule DSL (eq/ne/in/not_in/gt/gte/lt/lte/exists over dotted paths, no eval, §8). "
     "Node-N routing: idempotency guard -> match active rules -> deferred matches to DeferredSink "
     "(M06/A7), in-session matches to A6.reserve; approved -> FireMessage on FireSink (M08/B2), "
     "suppressed -> SuppressionSink (Z1/M14), no match -> dropped. Contracts are messages to sinks "
     "not direct calls (§5); FireMessage (generator/models.py) carries A6's reservation_id. "
     "RedisTriggerConsumer (XREADGROUP events:in/group trigger-eval -> evaluate -> XACK) + "
     "parse_stream_fields; end-to-end store->relay->parse->evaluate->fire tested with no Redis. "
     "Deferred (POA/04 §11): I/J derivation workers (need A7), rule hot-reload, handoff branch "
     "(§10.3), consumer partitioning.",
     "services/event_pipeline/triggers/{dsl,evaluator,consumer,sinks}.py, generator/models.py "
     "(FireMessage), POA/04 §11; tests/test_trigger_evaluation.py (12) — 146 total green, ruff "
     "clean. Commit 2b46310 (PR #1)"),

    ("A3", "Customer Journey & Behavioural Event Capture (SDK / contract)",
     "POA/01", "Track A", "Not started", "", "", "",
     "Depends on the A4 contract, not the whole service."),

    ("A7", "Pending-Engagement Queue & Deferred Scheduler (Celery)",
     "POA/06", "Track A", "In progress", "2026-09-02", "",
     "Core done + tested. services/event_pipeline/pending/: pending_engagements queue where "
     "PendingQueue.enqueue IS A5's DeferredSink (so A5's deferred matches flow straight in, no "
     "duplication). eligible_at = created + wait_period, expires_at = created + expiry; every read "
     "takes an explicit `now` so wait/eligibility/expiry are time-travel tested. expire_due sweeps "
     "overdue pending -> expired (returns ids for Z2), never touching raised/raising; reconcile_stuck "
     "releases crashed `raising` rows. Login re-eval (node S): on_login under a per-customer lock "
     "fetches eligible entries, arbitrates them through A6 (multiple pending compete on PRECEDENCE, "
     "§10.3), fires the winner via FireSink + marks raised; losers/cap-suppressed stay pending for a "
     "later login. Concurrency test: 5 parallel logins raise at most once. Celery Beat wiring "
     "(beat_schedule + register_pending_tasks) with no top-level celery import. e2e A5->A7->A6 "
     "tested. Deferred (POA/06 §11): in-session merge at login, SKIP-LOCKED claim, Beat integration run.",
     "services/event_pipeline/pending/{tables,queue,scheduler,tasks,bootstrap}.py, "
     "services/requirements.txt + pyproject (celery), POA/06 §11; tests/test_pending_queue.py (12) "
     "— 158 total green, ruff clean. Commit 10285b5 (PR #1)"),

    ("A8", "Human Handoff Manager",
     "POA/07", "Track A", "Not started", "", "", "",
     "NEXT. Depends on A5 (done). Keep handoff code THIN around RoutingRule .match/.route/.sla dicts "
     "— B1 will likely tighten them into real models. Handoff event is a message contract, not a "
     "call (§5). Routing defaults already in generator/fixtures.default_routing_rules()."),

    # --- Process / cross-track --------------------------------------------- #
    ("P1", "Accept Track A + maintain POA/18 §7 status rows",
     "POA/18", "Process", "DONE", "2026-09-01", "2026-09-01",
     "Accepted Track A ownership (event/trigger pipeline + platform, POA 01-07 & 15). Maintain my "
     "own A- and Sprint-0 rows in the shared POA/18 §7 status table. (POA/18 itself authored by "
     "Shagun.)",
     "POA/18 §7 (my rows)"),

    ("P3", "Ratify service-code repo layout (POA/18 §8.2)",
     "POA/18 §8.2", "Process", "DONE", "2026-09-01", "2026-09-01",
     "RESOLVED from my side: chose the services/ monorepo — services/platform (shared template), "
     "services/event_pipeline (Track A), services/conversation (Track B). Recorded it in POA/15 "
     "§12 and marked POA/18 §8.2 RESOLVED. This unblocks B1 scaffolding for Shagun.",
     "POA/15 §12, POA/18 §8.2. Commit 4a1ed3d"),

    ("P4", "Deterministic test_data — seed event ids + regenerate",
     "design principle P4", "Process", "DONE", "2026-09-01", "2026-09-01",
     "The committed test_data churned on every `generator build` because golden + volume event ids "
     "used uuid4. Replaced them with a seeded UUIDv5 (rng.stable_uuid) keyed on stable fields, "
     "proved two independent regens are byte-identical, then regenerated test_data as the canonical "
     "snapshot — which also wrote Shagun's S3 conversations/ output for the first time "
     "(deterministic, so a Track B regen reproduces it identically, no conflict).",
     "generator/rng.py, scenarios.py, patterns.py; test_data/*. Commits da93227, 2eeea2b"),

    ("P2", "Follow git workflow (POA/18 §6) + PR into main",
     "POA/18 §6", "Process", "In progress", "2026-09-01", "",
     "Following §6: long-lived branch track-a/event-pipeline, merge (not rebase) main in at "
     "day-end, PR into main (PR #1 open), no force-push. TO RECONCILE with Shagun: §6 describes "
     "per-person branches, but Shagun is currently keeping everything LOCAL/un-pushed — the two "
     "conventions need to agree.",
     "branch track-a/event-pipeline, PR #1"),
]

# --------------------------------------------------------------------------- #
# Sheet 2 — Work Log
# --------------------------------------------------------------------------- #
LOG_COLUMNS = [
    ("Date", 12),
    ("Task ID", 10),
    ("What I did", 96),
    ("Outcome / proof", 44),
    ("Pushed", 14),
]

LOG = [
    ("2026-08-27", "foundation",
     "Prior sessions (pre-split): built the generator/mocks/tests foundation and POA 00-17 — the "
     "seeded reference world, Pydantic contract models, 7 golden verification scenarios, and the "
     "v0.2 SME business domain (companies/rate-plans/invoices, booking lifecycle, van classes, "
     "protection/extras/policy catalogues) driven by the POA/17 audit. This is the shared "
     "foundation both tracks build on — the reason I own Track A.",
     "43 tests green; generator v0.2", "78c7d3b, 315ab9c"),

    ("2026-09-01", "S1/S2/S5",
     "Implemented Sprint 0 Track A. S1: 12-class taxonomy + city/suburban/US stations + one-way via "
     "an additive world pass-3 that keeps the golden prices byte-identical (verified against "
     "origin/main). S2: one-way/late-return/no-show/fuel fees + 5 pinned dispute fixtures, with "
     "post-rental charges kept OUT of the quoted total so disputes are realistic. S5: LoadProfile "
     "knobs from the §16.3 targets. Contract changes to models.py were additive/optional only.",
     "19 new tests, 92 total green", "73dde84"),

    ("2026-09-01", "P4",
     "Found the committed test_data churned on every regen because golden + volume event ids used "
     "uuid4. Replaced them with a seeded UUIDv5 keyed on stable fields, proved two independent "
     "regens are byte-identical, then regenerated test_data as the canonical snapshot (which also "
     "wrote Shagun's S3 conversations/ output for the first time — deterministic, no future conflict).",
     "regens diff-clean; 92 green", "da93227, 2eeea2b"),

    ("2026-09-01", "A1 / P3",
     "Ratified the services/ monorepo (POA/15 §12) and built the A1 platform skeleton: "
     "services/platform template (create_app wiring logging/correlation-id/metrics/OTel-seam/"
     "health/readyz/errors + lazy pg/redis/celery factories), a booting services/event_pipeline "
     "app, docker-compose (pg+redis+service), Dockerfile, GitHub Actions CI, .env.example. "
     "Installed the service runtime deps and kept the heavier infra deps lazy.",
     "8 new tests, 100 total green; ruff clean; /healthz /readyz /metrics all 200", "4a1ed3d (PR #1)"),

    ("2026-09-01", "tracker",
     "Reviewed Shagun's new 3-sheet tracker (Docs/Shagun_Task_Tracker.xlsx), my own tracker and "
     "POA/18, then rebuilt my Track A tracker to match Shagun's format (Task List / Work Log / "
     "Reference) via build_prasad_tracker.py.",
     "Docs/Prasad_TrackA_Task_Tracker.xlsx regenerated", "local"),

    ("2026-09-01", "A2",
     "Built the Event Store on the A1 platform template. Followed the project's delegate-to-the-"
     "spec discipline (persists generator.models.Event, so the generated dataset round-trips) and "
     "the transactional-OUTBOX pattern rather than a dual write — the whole exactly-once argument "
     "is publish-then-mark ordering. Made it testable without infrastructure the way Shagun did for "
     "B4: portable schema (JSONB/bigserial on Postgres, JSON/rowid on SQLite) + an injected stream "
     "publisher, so 12 tests incl. the §7 no-loss property run on in-memory SQLite with no Docker. "
     "One real bug the tests caught: a BigInteger outbox PK doesn't autoincrement on SQLite (only "
     "INTEGER PRIMARY KEY is the rowid alias) — fixed with with_variant. Scoped POA/03 §11 honestly: "
     "retention/partition + signal-J backfill + live-Redis deferred.",
     "12 new tests, 112 total green; ruff clean", "d8fc10b (PR #1)"),

    ("2026-09-01", "A4",
     "Built the Ingestion API — the single write-door — on the A1 template, writing through A2's "
     "outbox (so no direct dual write). Kept the same discipline: the body IS the Event contract, "
     "so validation + extra='forbid' doubles as the PII field allow-list, and idempotency is the "
     "store's not a second mechanism. The §10 open questions (auth mechanism, identity token vs "
     "body, volume) are handled behind seams — API-key auth, identity binding, in-memory rate "
     "limiter — the way Shagun handled B4/B5's open transport, so none of them block code. Batch "
     "endpoint does per-item validation for partial success. Added an injected-store path to "
     "build_app for tests and an end-to-end test proving API->store->outbox->relay->stream.",
     "11 new tests, 123 total green; ruff clean", "ee06657 (PR #1)"),

    ("2026-09-01", "A6",
     "Built the Frequency Cap & Precedence engine — the gate A5 calls before firing. Same "
     "delegate-to-the-spec discipline as A2/A4: cap accounting hands ledger timestamps to "
     "reference.would_fire, the exact sliding-window rule test_invariants already asserts, so the "
     "service can't quietly diverge from the spec. Precedence is deterministic and order-independent "
     "(weight -> specificity -> recency -> id). The judgement call was the reserve->confirm/rollback "
     "handshake: a reserved slot is only finalised when M08 delivers, and rolled back on failure, so "
     "a failed send doesn't burn a customer's cap — the cross-track contract I put in "
     "generator/models.py to reconcile with Shagun's local §5. Wrote the concurrency invariant test "
     "POA/05 §6 asks for (10 threads, cap 1 -> exactly one fires) using a per-customer lock seam.",
     "11 new tests, 134 total green; ruff clean", "7a5a0f8 (PR #1)"),

    ("2026-09-01", "A5",
     "Built the trigger engine — node N, the brain that ties A2, A6 and (eventually) M08 together. "
     "Kept A5 thin on purpose: matching is a small sandboxed field/op/value DSL (no eval), "
     "cap/precedence is delegated to A6, and every downstream is a message to a sink rather than a "
     "direct call (POA/18 §5) so Track B's M08 and my own A7 plug in without A5 changing. The fire "
     "decision is a shared FireMessage carrying A6's reservation_id, so the confirm/rollback loop "
     "reaches back to the cap. Added an idempotency guard so an at-least-once stream redelivery "
     "can't double-fire, and an end-to-end test that runs a real ingested event through "
     "store->relay->parse->evaluate->fire without needing Redis. Scoped POA/04 §11 honestly: the "
     "I/J derivation workers wait on A7, and the handoff branch on M07/§10.3.",
     "12 new tests, 146 total green; ruff clean", "2b46310 (PR #1)"),

    ("2026-09-02", "A7",
     "Built the deferred branch — the pending queue + login re-evaluation. The clean win was making "
     "PendingQueue implement A5's DeferredSink, so A5's deferred matches flow straight in with no "
     "new contract, and on_login just re-arbitrates them through the SAME A6 engine — so precedence "
     "and cap behave identically whether a signal fires in-session or after a login. Kept it "
     "testable without a clock or a broker: every read takes an explicit `now` (so expiry/eligibility "
     "are time-travelled) and the Celery tasks are thin wrappers over pure methods with no top-level "
     "celery import. Wrote the two invariants the POA cares about — raised-at-most-once under "
     "concurrent logins, and expiry never touching a raised row. Answered §10.3 (pending entries "
     "compete on precedence, not FIFO).",
     "12 new tests, 158 total green; ruff clean", "10285b5 (PR #1)"),
]

# --------------------------------------------------------------------------- #
# Sheet 3 — Reference
# --------------------------------------------------------------------------- #
REFERENCE = [
    ("THE SPLIT", ""),
    ("Prasad — Track A", "Event & trigger pipeline + platform: POA 01-07, 15. Sprint 0: S1, S2, S5."),
    ("Shagun — Track B", "Conversation, config & reporting: POA 08-14. Sprint 0: S3, S4, S6."),
    ("", ""),
    ("SHARED FILES — announce before pushing", ""),
    ("generator/models.py", "Both tracks need it. Append one bounded contiguous block; never rename "
                            "or remove a field the other track uses. My Sprint 0 added optional "
                            "fields/enums only — nothing renamed/removed."),
    ("generator/pipeline.py", "Both add writers. Keep the diff to contiguous appended lines."),
    ("tests/conftest.py", "Keep new fixtures in your own test module unless genuinely shared."),
    ("Golden prices are sacred", "Any world change must keep the 7 golden-scenario prices "
                                 "byte-stable — use the additive-pass pattern (see S1)."),
    ("", ""),
    ("DECIDED — from my (Track A) side", ""),
    ("Repo layout", "RATIFIED: services/ monorepo (platform + event_pipeline + conversation). "
                    "Recorded in POA/15 §12; POA/18 §8.2 marked RESOLVED."),
    ("Git workflow", "Track A on long-lived branch track-a/event-pipeline; merge (not rebase) main "
                     "at day-end; PR into main (PR #1 open); no force-push. To reconcile with "
                     "Shagun's current local-only rule."),
    ("", ""),
    ("OPEN — needs the client (POA/00 §7)", ""),
    ("LLM provider & data residency", "Which model, which region, Hertz constraints."),
    ("Booking API", "One service or several? Latency SLA? (gates A-side claim inputs / B4.)"),
    ("HS-103 surface", "REST, websocket or embedded widget SDK?"),
    ("Support/agent queue", "Zendesk, Salesforce or in-house? (gates A8 handoff routing.)"),
    ("", ""),
    ("HOW TO USE THIS FILE", ""),
    ("On starting a task", "Set Status to 'In progress' and fill Started."),
    ("On finishing", "Set Status 'DONE', fill Completed, and write What I did + Files/evidence."),
    ("Every session", "Add a Work Log row — the log is what shows effort, the status is only a flag."),
    ("Regenerate", "python build_prasad_tracker.py  (overwrites this file — edit the script, not "
                   "just the sheet, if you want changes to survive)"),
]


def _style_header(ws, columns, row=1):
    for i, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 28


def build() -> None:
    wb = Workbook()

    # ---- Sheet 1: Task List ---------------------------------------------- #
    ws = wb.active
    ws.title = "Task List"
    ws["A1"] = "Prasad — Track A task list (HFB Proactive AI Chatbot)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = ("Track A = event & trigger pipeline + platform. Shagun owns Track B (conversation, "
                "config & reporting). See POA/18 for the full split.")
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    ws.merge_cells("A2:I2")

    _style_header(ws, TASK_COLUMNS, row=4)
    for r, row in enumerate(TASKS, start=5):
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        ws.cell(row=r, column=5).fill = STATUS_FILL.get(row[4], TODO_FILL)
        ws.cell(row=r, column=5).font = Font(bold=True)
        ws.row_dimensions[r].height = 108 if row[4] == "DONE" else 40

    dv = DataValidation(
        type="list", formula1='"Not started,In progress,Blocked,DONE"', allow_blank=False
    )
    ws.add_data_validation(dv)
    dv.add(f"E5:E{4 + len(TASKS)}")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{4 + len(TASKS)}"

    # ---- Sheet 2: Work Log ------------------------------------------------ #
    ws2 = wb.create_sheet("Work Log")
    ws2["A1"] = "Work log — one row per working session"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:E1")
    ws2["A2"] = "Add a row every day. This is what shows the effort; Status on sheet 1 is only a flag."
    ws2["A2"].font = Font(italic=True, size=9, color="595959")
    ws2.merge_cells("A2:E2")

    _style_header(ws2, LOG_COLUMNS, row=4)
    for r, row in enumerate(LOG, start=5):
        for i, val in enumerate(row, start=1):
            c = ws2.cell(row=r, column=i, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        ws2.row_dimensions[r].height = 70
    ws2.freeze_panes = "A5"

    # ---- Sheet 3: Reference ----------------------------------------------- #
    ws3 = wb.create_sheet("Reference")
    ws3["A1"] = "Reference — split, shared files, decisions, open questions"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:B1")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 104
    for r, (k, v) in enumerate(REFERENCE, start=3):
        a = ws3.cell(row=r, column=1, value=k)
        b = ws3.cell(row=r, column=2, value=v)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if v == "" and k:                       # section heading
            a.font = Font(bold=True, color="1F3864")
            a.fill = PatternFill("solid", fgColor="D9E2F3")
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        else:
            a.font = Font(bold=True)

    wb.save(OUT)
    print(f"wrote {OUT}: {len(TASKS)} tasks, {len(LOG)} log rows")


if __name__ == "__main__":
    build()
