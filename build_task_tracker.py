"""Generate Shagun_Task_Tracker.xlsx — Shagun's (Track B) task list and work log.

Regenerate with:  python build_task_tracker.py [--open]

`--open` launches the sheet in the default application afterwards. Run it that
way at the end of every task: update the TASKS / LOG lists below, re-run, and
the sheet opens showing what changed.

Sheet 1 "Task List"  — every Track B / Sprint-0 task, its POA reference, status,
                       completion date and a "What I did" column filled in as
                       each task lands.
Sheet 2 "Work Log"   — dated, one row per session/commit, so the tracker shows
                       *what was actually done* rather than only a status flag.
Sheet 3 "Reference"  — track split, shared-file rules, open questions.
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "Shagun_Task_Tracker.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
PROG_FILL = PatternFill("solid", fgColor="FFEB9C")
TODO_FILL = PatternFill("solid", fgColor="F2F2F2")
BLOCK_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REVERT_FILL = PatternFill("solid", fgColor="E4DFEC")

STATUS_FILL = {
    "DONE": DONE_FILL,
    "In progress": PROG_FILL,
    "Not started": TODO_FILL,
    "Blocked": BLOCK_FILL,
    "Reverted": REVERT_FILL,
}

# --------------------------------------------------------------------------- #
# Sheet 1 — Task List
# --------------------------------------------------------------------------- #
TASK_COLUMNS = [
    ("Task ID", 10),
    ("Task", 46),
    ("POA reference", 24),
    ("Phase", 14),
    ("Status", 13),
    ("Started", 12),
    ("Completed", 12),
    ("What I did (fill in on completion)", 74),
    ("Files / evidence", 46),
]

TASKS = [
    # --- Sprint 0: shared test-dataset backlog (POA/16 §16) ----------------- #
    ("S3", "Conversation-intent scenarios + scripted trees (17 intents)",
     "POA/16 §16.4, §16.6", "Sprint 0", "DONE", "2026-09-01", "2026-09-01",
     "Built IntentScenarioComposer with one scripted tree per mandated intent — depth varied by "
     "structural distinctness (single-turn lookup, multi-turn slot filling, out-of-scope refusal, "
     "ambiguity->clarify, and CV-17 mid-conversation requirement change as a real branching tree). "
     "Shipped BOTH halves of the §16.6 Phase-2 seam: ReplySource protocol (ScriptedReplySource for "
     "Phase 1) and Evaluator protocol (ExactExpectationEvaluator, judges a transcript not the script "
     "so it survives LLM-generated replies). All claims grounded in the same World the booking-API "
     "mock reads. Split delivered_excludes (claim WRONG vs live data, must never be delivered) from "
     "superseded_tokens (claim CORRECT when said, then obsoleted — must not reach the final "
     "confirmation); conflating them made one of the two cases silently stop testing anything.",
     "generator/intents.py (new), models.py, pipeline.py, cli.py, "
     "tests/test_conversation_intents.py — 73 tests green. Reverted then restored "
     "2026-09-01; in the tree and green. Local only, not pushed."),

    ("S4", "PII redaction fixtures — obvious + embedded PII",
     "POA/16 §16.5", "Sprint 0", "DONE", "2026-09-01", "2026-09-01",
     "13 fixtures: 3 obvious (labelled/form-like), 7 embedded in natural conversational prose "
     "(the hard case §16.5 calls out), 3 edge cases — punctuation-adjacent PII, two similar-shaped "
     "identifiers that must not collapse into one pattern, and a PII-free negative control. All 12 "
     "PIIKind values covered. Put redact() in reference.py with the other trust-critical decisions "
     "rather than in pii.py, since redact-before-LLM is exactly that class of decision (M15 §4); "
     "it is offset-addressed like M10's claim resolution, so it never regex-guesses, and it raises "
     "on drifted or overlapping spans instead of mangling output. Spans are computed at build time "
     "via .find() with a uniqueness assert — hand-counted offsets drift the moment anyone edits a "
     "sentence. Every synthetic value is reserved BY SPECIFICATION and the tests assert the "
     "property not the format (cards fail Luhn, emails RFC 2606/6761, phones Ofcom 07700 900xxx "
     "drama block, IPs RFC 5737 TEST-NET-1), so a plausible real-looking value fails the suite. "
     "Added the over-redaction guard — every fixture carries a `preserves` list, because 'no PII "
     "survives' passes trivially for a redactor that destroys the whole message. PII_FIELDS marks "
     "the PII-bearing model fields with NOT_PII recording deliberate exclusions; a test asserts "
     "both still match the models, so the marking can't rot when Prasad's S1/S2 touch models.py.",
     "generator/pii.py (new), reference.py, models.py, pipeline.py, cli.py, "
     "tests/test_pii_redaction.py (26 tests) — 99 total green. "
     "Output: test_data/fixtures/pii_redaction.json. Local only, not pushed."),

    ("S6", "Future-client-compat layer — repository interface, field_map.yaml, lenient DTO",
     "POA/16 §16 item 6", "Sprint 0", "DONE", "2026-09-01", "2026-09-01",
     "Shipped the three named artifacts: ReferenceRepository protocol, generator/field_map.yaml, "
     "and the lenient coerce() DTO. Scoped the claim honestly rather than repeating the S3 "
     "overclaim: the spec says 'keeps 1-8 swappable', but a repository over entity data only "
     "covers §16.7 components 1-3. Components 4-8 were ALREADY swappable (GenConfig instance, "
     "PII_FIELDS dict, S3's ReplySource and Evaluator protocols) — so S6 adds the seam for 1-3 and "
     "documents the swap point for all eight in a table, which makes 'minimal code change' a "
     "checkable claim instead of a slogan. Derived the interface from real call sites (grepped "
     "every attribute reached through World) rather than inventing it, so it neither over- nor "
     "under-specifies. GeneratedRepository is a thin pass-through and a test asserts it agrees "
     "with the world on EVERY rate and availability key, not a spot-check — drift there would "
     "quietly break claim verification. Leniency is boundary-only: the contract models keep "
     "extra='forbid' and a test pins that S6 didn't relax it; coerce() reports every rename, "
     "value-map and drop in a CoercionReport because a silently dropped field is data loss, and "
     "strict=True refuses rather than drops. field_map.yaml sits beside the code, not under "
     "test_data/, so regenerating the dataset can't delete it. "
     "BUG CAUGHT BY TEST: the enum-target validator used issubclass(annotation, Enum), which "
     "matches nothing for optional fields like `Transmission | None` — so it was silently "
     "validating zero value-maps. Fixed with _enum_of() unwrapping the union.",
     "generator/repository.py (new), generator/field_map.yaml (new), README, POA/16, "
     "tests/test_repository_compat.py (22 tests) — 121 total green. Local only, not pushed."),

    # --- Track B: service modules ------------------------------------------ #
    ("B1", "Admin Console & Trigger Configuration — persistence + CRUD",
     "POA/13", "Track B", "DONE (service layer)", "2026-09-02", "2026-09-02",
     "Shipped task 3 fully and 1/2/4 in part — POA/13 §9 wants data model + CRUD early so M04/M05 "
     "have config, and that is what landed. Tasks 5/6/7 are genuinely blocked: RBAC needs M15's "
     "auth model (Prasad's A1) plus §10.3's role matrix, the UI needs §10.1 answered (SPA or "
     "embedded?), and dry-run needs the shared DSL validator. The FastAPI surface is deferred "
     "because it adds a dependency to a shared file while the services/ layout is still "
     "unconfirmed after six modules; the service layer is transport-agnostic so routes are "
     "additive later. THE DSL DECISION: TriggerMatch.conditions is untyped list[dict] and §8 says "
     "ONE validator shared with M04 — writing a second here manufactures exactly the divergence §8 "
     "warns about, so everything AROUND the DSL is validated and conditions go through a named "
     "DSLValidator seam whose default validates nothing AND SAYS SO (an advisory issue tells the "
     "admin it went unchecked). A hole with a shape beats a guess. Two properties a naive version "
     "gets wrong: rollback CREATES a version rather than deleting one (overwriting would punch a "
     "hole in the audit exactly where someone reverted a bad config), and the audit is immutable "
     "in practice — frozen entries, reads return a tuple, so history cannot be edited through a "
     "returned value. Instant disable bypasses validation because the emergency brake must work on "
     "a config that could no longer be published. Publishes a VERSION STAMP not a payload — a "
     "payload would bake in a wire format invented on Prasad's behalf. generator/fixtures.py is "
     "read never written (his file, POA/18 §2), and tests pin both that and that his shipped "
     "fixtures pass their own validator.",
     "services/admin/config/{models,validation,service}.py, "
     "tests/test_admin_config_service.py (38 tests) — 320 total green. "
     "Consolidated all 5 open Prasad questions into POA/18 §5b. Local only, not pushed."),

    ("B2", "Conversation Orchestrator (context + personalisation)",
     "POA/08", "Track B", "DONE", "2026-09-01", "2026-09-01",
     "All 7 §5 tasks. The module that wires M09->M10->M11 into one conversation and owns the "
     "state M12 continues. Biggest judgement call was the PII claim: S4 built redact() which "
     "APPLIES redaction but does not DETECT it, and M08 has no detector — inventing one would "
     "repeat the mistake S4 avoided. So the guarantee is the smaller checkable one matching "
     "POA/15 §4 field allow-lists: the bundle carries only allow-listed fields and a test "
     "asserts that set is DISJOINT from every field PII_FIELDS marks as PII, so adding a PII "
     "field fails the suite instead of shipping a name to a provider. FREE_TEXT_FIELDS is empty "
     "AND asserted empty, making free-text a failing test rather than a silent regression. This "
     "is the end-to-end PII assertion I explicitly scoped OUT of S4 as an M08/M09 acceptance "
     "test — it is now real. Five failure points each tested independently; the trap is M10 "
     "returning blocked=True, which is a SUCCESSFUL call yielding no deliverable text, so a "
     "naive orchestrator sends nothing. Delivery failure rolls the M05 reservation back — "
     "otherwise the customer silently loses a capped engagement they never received. Prompt "
     "injection is mitigated not solved, and the POA says so: the fence raises the cost, M10 is "
     "what actually holds, and there is a test proving an injected £1/day price still gets "
     "killed by verification. TWO BUGS CAUGHT: json.dumps was escaping non-ASCII so £ became a "
     "backslash-u escape — which silently voids any assertion about what reached the provider; "
     "and M09's off-scope heuristic was a false positive on price quotes ('It's £52.21/day.' "
     "has no on-scope keyword), replacing good replies with generic fallbacks.",
     "services/conversation/orchestrator/{context,personalisation,prompts,state,service}.py, "
     "tests/test_orchestrator_service.py (30 tests) — 282 total green. Added the reservation "
     "confirm/rollback handshake to POA/18 §5. Local only, not pushed."),

    ("B3", "LLM Integration & Fallback Service",
     "POA/09", "Track B", "DONE (7/7)", "2026-09-01", "2026-09-01",
     "ALL SEVEN §5 TASKS DONE. Real AnthropicProvider on claude-opus-5: claims come back as "
     "STRUCTURED OUTPUT via output_config.format rather than parsed from prose, which realises "
     "the M09->M10 contract POA/10 §3.1 recommended and gives M10's exact detection path real "
     "input. Answered §10.2 in the process: the API returns NO numeric confidence, so the schema "
     "asks the model to self-report one and the gate stays conservative because a self-report is "
     "weaker than a logprob. Runs at effort=low with small max_tokens since generation is inline "
     "before delivery; left thinking adaptive rather than disabling it, which has documented "
     "failure modes on Opus 5. A claim whose token is not in the text, or lacking route context, "
     "is DROPPED — an unverifiable tag looks like coverage while M10 has nothing to address. "
     "Malformed model output degrades to an empty draft (->fallback) rather than raising. Added "
     "BudgetGuard: three limits that fail differently (session tokens, customer-daily tokens, "
     "global daily SPEND in money); exceeding one returns the same safe fallback as an outage, "
     "never an error. Prices are config not constants — a stale rate silently under-reports. "
     "Original build: full Y->W->X path: provider adapter, availability/confidence gate, and the localised "
     "fallback catalogue. Same delegation principle as M10 — the W threshold calls "
     "reference.decide_llm, which GS-06 already asserts against, so the golden scenario keeps "
     "covering what ships; the refusal/off-scope/length heuristics layer around it rather than "
     "replacing it. Highest-value test in the module: NO FALLBACK TEMPLATE MAY ASSERT A PRICE OR "
     "AVAILABILITY, checked over all 8 signals x 4 locales — fallbacks fire when the pipeline is "
     "already degraded, so a template quoting a figure would walk straight around M10's "
     "verification. Also: a missing template slot degrades to the generic localised message rather "
     "than showing a customer 'options for {route}'; an unsupported locale gets English AND is "
     "flagged so the gap surfaces; fallbacks carry no claims by design. Extracted "
     "CircuitBreaker/TTLCache to services/common/resilience.py since M09 needed the same breaker "
     "as M10 — verified all 144 M10-era tests still passed BEFORE writing any M09 code, so a "
     "refactor bug couldn't masquerade as an M09 bug. Two deviations recorded in POA/09 §11: sync "
     "not async (async would pull pytest-asyncio into shared requirements-dev.txt), and retry "
     "jitter deferred (can't be asserted deterministically; a flaky test in a sub-second suite is "
     "worse than a documented gap).",
     "services/conversation/llm/{provider,fallback,service}.py, services/common/resilience.py, "
     "tests/test_llm_fallback_service.py (33 tests) — 177 total green. "
     "NOW 7/7. Added the real AnthropicProvider on claude-opus-5 and the BudgetGuard. "
     "Local only, not pushed."),

    ("B4", "Claim Verification Service",
     "POA/10", "Track B", "DONE (7/7)", "2026-09-01", "2026-09-01",
     "ALL SEVEN §5 TASKS DONE. Added HTTPBookingAPIClient with bearer/API-key/HMAC auth read "
     "from the environment, never literals, and __repr__ overridden on every auth type so a "
     "config in a stack trace cannot print the secret. §10.1 is still open so the endpoint shape "
     "is an explicit assumption confined to BookingAPIEndpoints plus two parse methods — a test "
     "proves a totally different endpoint shape works without touching the client. Transport is "
     "injected (stdlib urllib) so 500s/timeouts/malformed bodies are tested with no socket and "
     "no sleeping. Key asymmetry: a malformed 200 is an OUTAGE (counts to the breaker), a 404 is "
     "a BAD LOOKUP (does not) — different operational facts. Added TolerancePolicy with five "
     "modes so answering §10.2 is a config change; at_least ('from £X') is the one easy to get "
     "backwards — it holds when the live price is AT OR BELOW the quote and fails when higher, "
     "because a customer quoted 'from £42' and charged £55 was misled. Default stays strict: too "
     "tight corrects a correct price (harmless), too loose ships a wrong one (the whole failure "
     "mode). Every VerificationRecord now carries tolerance_rule. Original build: First real service module in the project. Full AA->AB->AC path: detection, booking-API edge "
     "(timeout, circuit breaker, short-TTL cache), resolution and audit logging. Key decision: "
     "resolution DELEGATES to reference.apply_verification rather than reimplementing it — that "
     "function is the executable spec and test_invariants/test_golden_scenarios already assert "
     "against it, so a second copy would mean those suites stopped covering what actually ships; a "
     "test pins that service and reference agree. Tagged claims are the primary detection path "
     "(POA/10 §3.1) with a pattern fallback whose GAPS ARE PINNED BY TEST (misses 'forty-two "
     "pounds' and unsymboled amounts) — §8's first risk is detector brittleness and a fallback "
     "that quietly under-detects is worse than none. Outage vs no-data are distinguished in the "
     "log via FailureKind without widening the shared VerifyStatus enum; a bad lookup does not "
     "trip the breaker (would open the circuit on healthy infra); failures are never cached (a "
     "cached outage would poison the whole TTL). A draft quoting money with no verifiable context "
     "is refused rather than delivered. BUG CAUGHT BY TEST: the §3.3 guardrail used substring "
     "containment, but correcting 'available' -> 'not currently available' means the replacement "
     "CONTAINS the token — flagging a correct rewrite as a failure. Now counts occurrences. "
     "Scoped honestly in POA/10 §11: 5 of 7 §5 tasks done; real HTTP client + auth deferred "
     "pending the §10.1 endpoint answers, tolerance policy pending §10.2.",
     "services/conversation/claim_verification/{detection,client,service}.py, "
     "tests/test_claim_verification_service.py (23 tests) — 144 total green. "
     "NOTE: services/ layout is POA/18's unconfirmed assumption; tree kept shallow so a rename is "
     "one git mv. Local only, not pushed."),

    ("B5", "Chatbot UI Integration (HS-103) + admin UI",
     "POA/11", "Track B", "DONE", "2026-09-01", "2026-09-01",
     "All 7 §5 tasks implemented (task 1 is a conversation with the HS-103 team, not code). "
     "Transport is a protocol because §10.1 is open — §8's mitigation for 'HS-103 capabilities "
     "unknown' is adapter abstraction, so presence, anti-nag, correlation, receipts and retry all "
     "sit ABOVE the transport and do not move when the real surface lands. Anti-nag here is "
     "deliberately NARROWER than M05's frequency cap: re-implementing that would double-count and "
     "silently halve the configured cap, so this only blocks stacking a second proactive message "
     "on an unacknowledged conversation, and never blocks replies inside a live exchange (which "
     "would break the M12 loop). BUG CAUGHT BY TEST: I conflated delivery_ref (per MESSAGE) with "
     "thread_id (per CONVERSATION), so a second message tried to rebind the conversation — that "
     "would have misrouted every subsequent reply. Inbound is idempotent and never raises: a "
     "webhook delivering twice is normal and a 500 back just makes HS-103 retry the same "
     "unusable event. A failed delivery binds no correlation. Losing the button must not lose "
     "the message.",
     "services/conversation/delivery/service.py, tests/test_delivery_service.py (23 tests) — "
     "252 total green. Local only, not pushed."),

    ("B6", "Customer Response & Multi-turn Conversation Manager",
     "POA/12", "Track B", "DONE (5.5/7)", "2026-09-02", "2026-09-02",
     "The AE-AK loop: response detection, multi-turn with retained context, resolution/stuck "
     "detection, deep link, attribution, handoff raising, terminal logging. Tasks 5 and 6 are "
     "partial for reasons outside my control — AJ has no real booking-signal producer (M01/A3 "
     "does not exist) and the handoff contract needs agreeing with Prasad. "
     "FINDING WORTH THE MOST: reference.terminal_state has THREE outcomes (no_engagement / "
     "converted / handed_off) but the flow has FOUR — a customer the bot helped, shown a deep "
     "link, who has not yet booked is none of them. Calling terminal_state(True, True) at "
     "resolution returns `converted` and COUNTS A BOOKING THAT NEVER HAPPENED, inflating the exact "
     "conversion metric M14 reports and this feature is judged on. A test caught it. M12 leaves "
     "terminal as None between AI and AJ and surfaces the gap rather than forking the spec — "
     "POA/18 §5b item 7. Other decisions: max-turns ESCALATES rather than closing (a customer "
     "dropped mid-conversation is worse than one handed to a person); the handoff routes through "
     "M04 not straight to M07, because M04 is the single place that decides what happens to a "
     "signal; the event carries no queue/agent/skill — raised, never handled inline — but does "
     "carry the whole transcript, since a handoff that makes the customer repeat themselves is the "
     "failure this feature exists to prevent; `resolved` has to be earned (default unresolved, "
     "only explicit confirmation reaches it) because erring toward a person costs an agent's time "
     "while erring the other way strands someone; complaints and billing disputes are never "
     "bot-resolvable whatever the customer says, which is what the S3 CV-12/CV-13 trees pinned; "
     "and a booking BEFORE the conversation is never attributed, with the window boundary "
     "inclusive and asserted at both edges.",
     "services/conversation/response/{service,resolution,handoff}.py, "
     "tests/test_response_manager.py (31 tests) — 437 total green. "
     "Added POA/18 §5b items 6 and 7. Local only, not pushed."),

    ("B7", "Audit, Reporting & Analytics",
     "POA/14", "Track B", "DONE (6/7)", "2026-09-02", "2026-09-02",
     "Outcome-event contracts, immutable store, metric rollups, attribution join, query layer, "
     "CSV export and read-only views over M13's config audit. Task 6 (dashboard) deferred — "
     "frontend stack, §10.3 unanswered. Adapters mean NOTHING UPSTREAM HAD TO CHANGE to support "
     "reporting. THE THING THAT MATTERS MOST: every rate names its denominator, because that is "
     "where a reporting module gets silently wrong and the four in §3.2 do not share one — "
     "conversion is over FIRED not delivered, because an engagement approved and then undelivered "
     "is a conversion we LOST and moving it into the denominator would flatter the number. A zero "
     "denominator returns None, never 0.0: '0% conversion' and 'no data' are different facts and "
     "a dashboard that renders them identically lies to whoever decides from it. Aggregates are "
     "RECOMPUTED from raw events on every call, never incremented — a cached counter that drifts "
     "from its source is the classic reporting bug and it drifts silently; §6's reconciliation is "
     "asserted as a property over every event kind rather than a spot-check. M12's terminal=None "
     "finding flows through as a real PENDING bucket, with converted+pending+not_converted == "
     "conversations asserted and pending clamped at zero (a conversion can land in a later window "
     "than its resolution). The funnel is asserted never to widen. PII: aggregates and exports are "
     "PII-free BY CONSTRUCTION (no OutcomeEvent field is PII-marked, and no S4 fixture value "
     "reaches a CSV) — but drill-down into conversations reaches customer text and needs M15 "
     "access control, which does not exist; that boundary is stated rather than glossed.",
     "services/analytics/{events,metrics,service}.py, tests/test_analytics_service.py "
     "(35 tests) — 472 total green. Added POA/18 §5b item 8 (nobody emits Z1/Z2 to M14). "
     "Local only, not pushed."),

    # --- Process ------------------------------------------------------------ #
    ("P1", "Two-person work split + daily git workflow (POA/18)",
     "POA/18", "Process", "DONE", "2026-08-31", "2026-09-01",
     "Wrote POA/18: split the module dependency graph into two tracks so no module gets built "
     "twice, mapped cross-track contract points, and defined the daily push/merge routine. "
     "Corrected an early error in it — B1 does NOT block Prasad's A5/A6/A8, because the M13 "
     "contract already exists. Confirmed track ownership after Prasad accepted Track A.",
     "POA/18_Team_Work_Split_and_Git_Workflow.md + POA/00 index row. Reverted then restored "
     "2026-09-01; in the tree. Local only, not pushed."),

    ("P2", "Agree git workflow (POA/18 §6) with Prasad",
     "POA/18 §6", "Process", "Blocked", "", "",
     "", "Waiting on Prasad. §6 still describes per-person branches while we push to main."),

    ("P3", "Agree service-code repo layout (services/*) with Prasad",
     "POA/18 §8.2", "Process", "Blocked", "", "",
     "", "Waiting on Prasad. Blocks B1 scaffolding. No POA specifies a layout yet."),
]

# --------------------------------------------------------------------------- #
# Sheet 2 — Work Log
# --------------------------------------------------------------------------- #
LOG_COLUMNS = [
    ("Date", 12),
    ("Task ID", 10),
    ("What I did", 96),
    ("Outcome / proof", 44),
    ("Pushed", 12),
]

LOG = [
    ("2026-08-31", "P1",
     "Read the whole repo and all 18 POA files. Wrote POA/18: cut the module dependency graph into "
     "two tracks that live in separate directories, so parallel work merges additively instead of "
     "colliding. Added Sprint-0 file-level split, cross-track contract points and a status tracker.",
     "POA/18 created, indexed in POA/00", "d1f6987"),

    ("2026-09-01", "P1",
     "Prasad accepted Track A. Marked ownership CONFIRMED in POA/18 and closed open question 1.",
     "POA/18 v1.1", "3f444b8"),

    ("2026-09-01", "S3",
     "Built the 17 scripted conversation trees, the ReplySource seam, world-grounded claims, and "
     "the pipeline/CLI wiring to write them to test_data/conversations/.",
     "23 new tests, 66 total green", "9840cb0"),

    ("2026-09-01", "S3",
     "Fixed a contract bug found in review: delivered_excludes was carrying two incompatible "
     "meanings. Split out superseded_tokens, added the Evaluator half of the §16.6 Phase-2 seam, "
     "promoted path()/paths()/all_turns() so there is only one definition of a path, and extended "
     "the tree-wide invariants to cover branch turns (they were escaping world-grounding).",
     "73 tests green", "78aedf0"),

    ("2026-09-01", "—",
     "Reverted all five commits (POA/18 + S3 + tracker) at Shagun's request. Used git revert "
     "rather than a force-push so history stays intact and Prasad's clone is unaffected. Working "
     "tree verified byte-identical to 67b977c; 43 tests green. New standing rule from here: all "
     "work stays LOCAL, nothing goes to GitHub unless explicitly asked.",
     "Tree == 67b977c, 43 tests green", "local only"),

    ("2026-09-01", "S3 / P1",
     "Restored both reverted deliverables locally at Shagun's request — S3 first, then POA/18. "
     "Restored at file level from 78aedf0 / b54edb8 rather than by reverting the revert, so each "
     "came back independently and history stays readable. Restoring POA/18 also re-validated the "
     "'see POA/18 §4' pointer in generator/models.py, which was dangling in between.",
     "Working tree == b54edb8 (plus tracker improvements); 73 tests green", "local only"),

    ("2026-09-01", "S4",
     "Built the PII redaction fixture set — 13 fixtures across obvious / embedded / edge cases, "
     "covering all 12 PII kinds, plus the PII_FIELDS data dictionary and reference.redact(). "
     "Two decisions worth recording: redact() went in reference.py (not pii.py) because "
     "redact-before-LLM is a trust-critical decision like M10's claim resolution and belongs where "
     "a service author will look for it; and spans are computed at build time rather than "
     "hand-counted, so editing a fixture sentence can't silently break its offsets. Scoped the "
     "claim honestly in POA/16: this proves the redaction decision and the fixtures — 'no PII ever "
     "reaches the LLM' is an M08/M09 acceptance test, since no orchestrator exists yet to build a "
     "prompt.",
     "26 new tests, 99 total green; test_data/fixtures/pii_redaction.json", "local only"),

    ("2026-09-01", "S6",
     "Built the future-client-compatibility layer, closing Sprint 0 on my side. The judgement call "
     "was scope: the spec's 'keeps 1-8 swappable' is broader than a repository can deliver, so I "
     "shipped the seam for components 1-3 and documented the existing swap point for 4-8 in a "
     "table rather than claiming coverage I didn't have. Derived the protocol by grepping actual "
     "World call sites instead of inventing a method set. A test I wrote to catch typo'd enum "
     "targets immediately caught a real bug in my own validator — it used issubclass(annotation, "
     "Enum), which matches nothing for optional enum fields, so it had been validating zero "
     "value-maps.",
     "22 new tests, 121 total green; generator/repository.py + field_map.yaml", "local only"),

    ("2026-09-01", "B4",
     "Built M10 Claim Verification — the first real service module in the project, and the "
     "system's core trust guarantee. Biggest decision was NOT writing code: resolution delegates "
     "to reference.apply_verification, so the existing invariant and golden-scenario suites keep "
     "covering the shipped path instead of a divergent copy. Wrote the red-team test POA/10 §7 "
     "asks for — it tries to get an unverified price through every branch (correct/wrong/outage/"
     "timeout) and asserts none survives. A test caught a real bug in my own guardrail: it used "
     "substring containment, but correcting 'available' to 'not currently available' means the "
     "replacement contains the original token, so a correct rewrite was reported as a failure. "
     "Now counts occurrences. Scoped POA/10 §11 honestly — 5 of 7 tasks done, the HTTP client and "
     "tolerance policy genuinely blocked on the §10 client questions.",
     "23 new tests, 144 total green; first services/ module", "local only"),

    ("2026-09-01", "B3",
     "Built M09 LLM Integration & Fallback. Did the shared-resilience refactor FIRST and confirmed "
     "all 144 existing tests still passed before writing a line of M09 — otherwise a refactor bug "
     "and an M09 bug would have looked identical. Same delegation discipline as M10: the "
     "confidence threshold calls reference.decide_llm rather than reimplementing it. The test I "
     "care most about asserts no fallback template can assert a price or availability, across "
     "every signal and locale — fallbacks fire exactly when the pipeline is degraded, so that is "
     "the path where an unverified claim would slip past M10. Recorded two honest deviations in "
     "POA/09 §11 rather than quietly diverging from the spec: sync instead of async, and jitter "
     "deferred.",
     "33 new tests, 177 total green; services/conversation/llm/", "local only"),

    ("2026-09-01", "B3 / B4 / B5",
     "Took M09 and M10 from 5/7 to 7/7 and shipped M11. Loaded the claude-api reference before "
     "writing the Anthropic adapter rather than working from memory — worth it: the current API "
     "differs from what I'd have written (adaptive thinking, output_config.effort, no "
     "budget_tokens, structured outputs via output_config.format). The structured-output work is "
     "the real win: M09 now emits tagged claims, which is the contract POA/10 §3.1 asked for and "
     "makes M10's exact-detection path real instead of aspirational. For the two genuinely "
     "product-owned questions (§10.1 endpoints, §10.2 tolerance) I implemented every plausible "
     "answer behind config seams, so the decision no longer blocks code. Two bugs caught by tests I "
     "wrote for the purpose: the tolerance at_least direction, and conflating delivery_ref with "
     "thread_id in M11 — the latter would have misrouted replies in production.",
     "75 new tests, 252 total green; POA 09/10 at 7/7, POA 11 done", "local only"),

    ("2026-09-01", "B2",
     "Built M08, the orchestrator — the module that wires 09/10/11 into one conversation. Most of "
     "the value was in deciding what NOT to claim: the PII guarantee is a field allow-list "
     "asserted disjoint from PII_FIELDS, not a detector I'd have had to hedge about, and it "
     "finally makes real the end-to-end PII test I scoped out of S4. Same honesty on prompt "
     "injection — the fence is a mitigation, M10 is what holds, and there's a test proving an "
     "injected £1/day price still dies at verification. Found two bugs while testing: json.dumps "
     "escaping non-ASCII (which would have quietly voided the PII assertions), and an off-scope "
     "false positive in M09 that was replacing good price replies with generic fallbacks. Added "
     "the M05 reservation confirm/rollback handshake to POA/18 §5 — it's cross-track and needs "
     "ten minutes with Prasad.",
     "30 new tests, 282 total green; POA/08 done", "local only"),

    ("2026-09-02", "B1",
     "Built the M13 service layer — the part POA/13 §9 wants early so M04/M05 have config. The "
     "interesting decision was the rule DSL: §8 says ONE validator shared with M04, so rather than "
     "write a second one I validated everything around it and left `conditions` behind a named "
     "seam that reports it went unchecked. Also spent time on two properties a naive version gets "
     "wrong — rollback creating a version instead of deleting one, and an audit that is actually "
     "immutable rather than just append-only. Finally consolidated all five open questions for "
     "Prasad into POA/18 §5b: they have been raised piecemeal across six modules without landing, "
     "and they are answerable in one sitting.",
     "38 new tests, 320 total green; POA/13 service layer done", "local only"),

    ("2026-09-02", "B6",
     "Built M12, closing the post-delivery loop. The most valuable thing to come out of it was "
     "not code: a test caught that reference.terminal_state cannot express this flow — it has "
     "three outcomes where the diagram has four, and using it at the resolution point would have "
     "counted bookings that never happened, quietly inflating the conversion metric. Left it as "
     "None and surfaced the gap rather than forking the shared spec. Also merged Prasad's branch "
     "earlier today (393 green), then found and fixed three real defects in his engagement path "
     "— including a cap race that let 4 reservations through a cap of 1 under concurrency.",
     "31 new tests, 437 total green; POA/12 at 5.5/7", "local only"),

    ("2026-09-02", "B7",
     "Built M14, the last module with real scope. Spent most of the thinking on denominators "
     "rather than code: conversion over fired vs delivered changes the headline number, and a "
     "zero denominator rendering as 0% instead of no-data is the kind of bug that survives to a "
     "board slide. Made aggregates recount from raw on every call so §6's reconciliation is true "
     "by construction rather than by discipline. M12's terminal=None finding paid off here — it "
     "became a visible pending bucket instead of quietly under-reporting conversion. Reused the "
     "S4 PII trick a third time, and scoped it honestly: aggregates are safe, drill-down is not, "
     "and that needs M15.",
     "35 new tests, 472 total green; POA/14 at 6/7", "local only"),
]

# --------------------------------------------------------------------------- #
# Sheet 3 — Reference
# --------------------------------------------------------------------------- #
REFERENCE = [
    ("THE SPLIT", ""),
    ("Prasad — Track A", "Event & trigger pipeline + platform: POA 01-07, 15. Sprint 0: S1, S2, S5."),
    ("Shagun — Track B", "Conversation, config & reporting: POA 08-14. Sprint 0: S3, S4, S6."),
    ("", ""),
    ("SHARED FILES — announce before pushing", ""),
    ("generator/models.py", "Both tracks need it. Append one bounded contiguous block; never rename "
                            "or remove a field the other track uses."),
    ("generator/pipeline.py", "Both add writers. Keep the diff to contiguous appended lines."),
    ("tests/conftest.py", "Keep new fixtures in your own test module unless genuinely shared."),
    ("", ""),
    ("OPEN — waiting on Prasad", ""),
    ("Git workflow (POA/18 §6)", "Doc says per-person branches; we are pushing to main. Needs a decision."),
    ("Repo layout", "services/platform, services/event_pipeline, services/conversation — unconfirmed."),
    ("", ""),
    ("OPEN — needs the client (POA/00 §7)", ""),
    ("LLM provider & data residency", "Which model, which region, Hertz constraints."),
    ("Booking API", "One service or several? Latency SLA?"),
    ("HS-103 surface", "REST, websocket or embedded widget SDK?"),
    ("Support/agent queue", "Zendesk, Salesforce or in-house?"),
    ("", ""),
    ("WORKING RULES", ""),
    ("Everything stays local", "No pushes to GitHub. Commit locally only, unless Shagun explicitly "
                              "asks in that moment. origin/main is still at b54edb8; the revert "
                              "commit is local and un-pushed."),
    ("Sheet updates automatically", "After every completed task: update this sheet and open it. "
                                    "No need to ask."),
    ("", ""),
    ("HOW TO USE THIS FILE", ""),
    ("On starting a task", "Set Status to 'In progress' and fill Started."),
    ("On finishing", "Set Status 'DONE', fill Completed, and write What I did + Files/evidence."),
    ("Every session", "Add a Work Log row — the log is what shows effort, the status is only a flag."),
    ("Regenerate", "python build_task_tracker.py  (overwrites this file — edit the script, not just "
                   "the sheet, if you want changes to survive)"),
]


def status_fill(status: str):
    """Colour by status PREFIX, so qualified values still read correctly.

    A status is often more informative with a qualifier — "DONE (7/7)",
    "DONE (service layer)" — but an exact-match lookup silently falls through to
    the grey not-started fill, which then says the opposite of the text beside
    it. Matching the longest prefix keeps the colour and the words agreeing.
    """
    for key in sorted(STATUS_FILL, key=len, reverse=True):
        if status.startswith(key):
            return STATUS_FILL[key]
    return TODO_FILL


def _style_header(ws, columns, row=1):
    for i, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 28


def build() -> None:
    wb = Workbook()

    # ---- Sheet 1: Task List ---------------------------------------------- #
    ws = wb.active
    ws.title = "Task List"
    ws["A1"] = "Shagun — Track B task list (HFB Proactive AI Chatbot)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = ("Track B = conversation, config & reporting. Prasad owns Track A (event/trigger "
                "pipeline + platform). See POA/18 for the full split.")
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    ws.merge_cells("A2:I2")

    _style_header(ws, TASK_COLUMNS, row=4)
    for r, row in enumerate(TASKS, start=5):
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        ws.cell(row=r, column=5).fill = status_fill(row[4])
        ws.cell(row=r, column=5).font = Font(bold=True)
        ws.row_dimensions[r].height = 92 if row[4] == "DONE" else 34

    # `showErrorMessage=False` because several rows carry a qualified status
    # ("DONE (7/7)"). The dropdown is a convenience for editing by hand; it must
    # not flag the values the generator itself writes as invalid.
    dv = DataValidation(
        type="list", formula1='"Not started,In progress,Blocked,DONE,Reverted"',
        allow_blank=False, showErrorMessage=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"E5:E{4 + len(TASKS)}")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{4 + len(TASKS)}"

    # ---- Sheet 2: Work Log ------------------------------------------------ #
    ws2 = wb.create_sheet("Work Log")
    ws2["A1"] = "Work log — one row per working session"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:E1")
    ws2["A2"] = "Add a row every day. This is what shows the effort; Status on sheet 1 is only a flag."
    ws2["A2"].font = Font(italic=True, size=9, color="595959")
    ws2.merge_cells("A2:E2")

    _style_header(ws2, LOG_COLUMNS, row=4)
    for r, row in enumerate(LOG, start=5):
        for i, val in enumerate(row, start=1):
            c = ws2.cell(row=r, column=i, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        ws2.row_dimensions[r].height = 56
    ws2.freeze_panes = "A5"

    # ---- Sheet 3: Reference ----------------------------------------------- #
    ws3 = wb.create_sheet("Reference")
    ws3["A1"] = "Reference — split, shared files, open questions"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:B1")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 104
    for r, (k, v) in enumerate(REFERENCE, start=3):
        a = ws3.cell(row=r, column=1, value=k)
        b = ws3.cell(row=r, column=2, value=v)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if v == "" and k:                       # section heading
            a.font = Font(bold=True, color="1F3864")
            a.fill = PatternFill("solid", fgColor="D9E2F3")
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        else:
            a.font = Font(bold=True)

    wb.save(OUT)
    print(f"wrote {OUT}: {len(TASKS)} tasks, {len(LOG)} log rows")


def open_sheet() -> None:
    """Open the sheet in the default application (Windows/macOS/Linux)."""
    p = Path(OUT).resolve()
    try:
        if sys.platform == "win32":
            os.startfile(p)  # noqa: S606 — intended: hand the file to Excel
        elif sys.platform == "darwin":
            subprocess.run(["open", str(p)], check=False)
        else:
            subprocess.run(["xdg-open", str(p)], check=False)
        print(f"opened {p}")
    except OSError as exc:                       # no GUI / no handler registered
        print(f"could not open {p}: {exc}")


if __name__ == "__main__":
    build()
    if "--open" in sys.argv:
        open_sheet()
